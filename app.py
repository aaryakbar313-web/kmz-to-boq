import streamlit as st
import pandas as pd
import openpyxl
from zipfile import ZipFile
import xml.etree.ElementTree as ET
import io
import math

st.set_page_config(page_title="KMZ to BOQ Converter", layout="wide", page_icon="⚡")

st.markdown("""
    <style>
    .main-title { font-size:32px; font-weight:bold; color:#1E3A8A; margin-bottom:10px; }
    .sub-title { font-size:16px; color:#4B5563; margin-bottom:30px; }
    </style>
    <div class="main-title">⚡ FTTH KMZ to BOQ Auto-Injector</div>
    <div class="sub-title">Situs otomatis untuk menghitung material KMZ Google Earth dan menyuntikkannya ke template Excel BOQ Jaringan.</div>
""", unsafe_allow_html=True)

def hitung_jarak(coord1, coord2):
    try:
        lon1, lat1 = map(float, coord1.split(',')[:2])
        lon2, lat2 = map(float, coord2.split(',')[:2])
        R = 6371000 
        phi1, phi2 = math.radians(lat1), math.radians(lat2)
        dphi = math.radians(lat2 - lat1)
        dlam = math.radians(lon2 - lon1)
        a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2) * math.sin(dlam/2)**2
        return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    except:
        return 0

def parse_kmz(file_bytes):
    try:
        with ZipFile(io.BytesIO(file_bytes)) as z:
            kml_filename = [f for f in z.namelist() if f.endswith('.kml')][0]
            kml_content = z.read(kml_filename)
    except Exception:
        kml_content = file_bytes

    root = ET.fromstring(kml_content)
    namespaces = {'kml': 'http://www.opengis.net/kml/2.2'}
    summary_data = {'Line': {}, 'Point': {}}
    
    for placemark in root.findall('.//kml:Placemark', namespaces):
        name_node = placemark.find('kml:name', namespaces)
        name = name_node.text.strip() if name_node is not None else "Unknown"
        
        if placemark.find('.//kml:LineString', namespaces) is not None:
            coord_node = placemark.find('.//kml:coordinates', namespaces)
            if coord_node is not None:
                coords = coord_node.text.strip().split()
                total_panjang = 0
                for i in range(len(coords) - 1):
                    total_panjang += hitung_jarak(coords[i], coords[i+1])
                summary_data['Line'][name] = summary_data['Line'].get(name, 0) + round(total_panjang)
                
        elif placemark.find('.//kml:Point', namespaces) is not None:
            summary_data['Point'][name] = summary_data['Point'].get(name, 0) + 1
            
    return summary_data

col_up1, col_up2 = st.columns(2)
with col_up1:
    uploaded_kmz = st.file_uploader("1. Upload File Jaringan (.kmz atau .kml)", type=["kmz", "kml"])
with col_up2:
    uploaded_excel = st.file_uploader("2. Upload Template Excel BOQ Anda", type=["xlsx"])

if uploaded_kmz and uploaded_excel:
    st.toast("Files uploaded successfully!", icon="🔥")
    kmz_data = parse_kmz(uploaded_kmz.read())
    
    st.write("### 🔍 Hasil Analisis Objek Google Earth")
    c1, c2 = st.columns(2)
    with c1:
        st.info("**Kabel Terdeteksi (Meter):**")
        if kmz_data['Line']:
            st.dataframe(pd.DataFrame(kmz_data['Line'].items(), columns=['Nama Layer KMZ', 'Panjang (m)']))
        else:
            st.warning("Tidak ditemukan objek garis di KMZ.")
    with c2:
        st.info("**Perangkat / Tiang Terdeteksi (Unit):**")
        if kmz_data['Point']:
            st.dataframe(pd.DataFrame(kmz_data['Point'].items(), columns=['Nama Layer KMZ', 'Jumlah (Pcs)']))
        else:
            st.warning("Tidak ditemukan objek titik di KMZ.")
            
    wb = openpyxl.load_workbook(io.BytesIO(uploaded_excel.read()))
    updated_items = []
    
    for sheet in wb.worksheets:
        for row in range(1, sheet.max_row + 1):
            cell_desc = sheet.cell(row=row, column=2).value
            if cell_desc and isinstance(cell_desc, str):
                for item_name, qty in kmz_data['Line'].items():
                    if item_name.lower() in cell_desc.lower():
                        sheet.cell(row=row, column=4).value = qty
                        updated_items.append({"Sheet": sheet.title, "Item": cell_desc, "Qty": f"{qty} m"})
                for item_name, qty in kmz_data['Point'].items():
                    if item_name.lower() in cell_desc.lower():
                        sheet.cell(row=row, column=4).value = qty
                        updated_items.append({"Sheet": sheet.title, "Item": cell_desc, "Qty": f"{qty} pcs"})
                        
    if updated_items:
        st.success(f"Berhasil mencocokkan otomatis {len(updated_items)} item material!")
        with st.expander("Lihat detail item Excel yang berhasil terisi otomatis:"):
            st.table(updated_items)
            
        output = io.BytesIO()
        wb.save(output)
        st.markdown("---")
        st.download_button(
            label="📥 DOWNLOAD EXCEL BOQ YANG SUDAH TERISI",
            data=output.getvalue(),
            file_name="Hasil_Auto_BOQ.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.error("Sistem tidak menemukan kecocokan nama antara Layer KMZ dengan kolom deskripsi di Excel.")
